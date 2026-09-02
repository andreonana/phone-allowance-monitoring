"""
Paramètres applicatifs (seuils d'alerte configurables) persistés dans
`data/config.xlsx` (feuille CONFIG, format clé/valeur). Comme le reste de
l'application, aucune base de données : un simple classeur Excel.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

import openpyxl

DEFAULTS: dict[str, float] = {
    # Seuil de déclenchement de l'alerte "forte augmentation", en % de variation
    # (positive) par rapport au mois précédent. Cf. cahier des charges §10.
    "SEUIL_FORTE_HAUSSE_PCT": 50.0,
    # Seuil de déclenchement de l'alerte "forte consommation", en FCFA, sur le
    # total consolidé (Orange + MTN) d'un collaborateur pour le mois courant.
    "SEUIL_FORTE_CONSOMMATION_FCFA": 100000.0,
}


class ConfigRepository:
    def __init__(self, chemin_fichier: Path):
        self.chemin_fichier = Path(chemin_fichier)
        self.valeurs: dict[str, float] = dict(DEFAULTS)
        self._charger()

    def _charger(self) -> None:
        if not self.chemin_fichier.exists():
            return
        wb = openpyxl.load_workbook(self.chemin_fichier, data_only=True)
        if "CONFIG" not in wb.sheetnames:
            return
        ws = wb["CONFIG"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            cle, valeur = row[0], row[1]
            try:
                self.valeurs[str(cle)] = float(valeur)
            except (TypeError, ValueError):
                continue

    def save(self) -> None:
        self.chemin_fichier.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CONFIG"
        ws.append(["CLE", "VALEUR"])
        for cle, valeur in self.valeurs.items():
            ws.append([cle, valeur])

        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=str(self.chemin_fichier.parent))
        os.close(fd)
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, self.chemin_fichier)
        except Exception:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            raise

    def get(self, cle: str) -> float:
        return self.valeurs.get(cle, DEFAULTS.get(cle, 0.0))

    def set(self, cle: str, valeur: float) -> None:
        self.valeurs[cle] = float(valeur)
