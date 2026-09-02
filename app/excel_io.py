"""
Lecture Excel robuste des fichiers opérateurs (Orange / MTN).

Objectif (cahier des charges §13, §16) : reconnaître les colonnes utiles
(numéro, montant, éventuellement matricule/staff id, éventuellement date)
quel que soit :
    - leur ordre dans le fichier,
    - le libellé exact de l'en-tête (tant qu'il reste sémantiquement proche),
    - la présence d'espaces parasites dans les noms de feuille/colonne,
    - la présence de colonnes supplémentaires non attendues.

Le fichier n'est JAMAIS filtré/tronqué de façon silencieuse : toute ligne
sans montant exploitable ou sans numéro est conservée et remontée avec un
statut explicite par le moteur de rapprochement (app.reconciliation_engine),
pas ici.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Optional

import openpyxl
import pandas as pd


def normalize_header(value) -> str:
    """Normalise un en-tête de colonne ou un nom de feuille pour comparaison
    tolérante : trim, espaces multiples réduits, casse uniformisée."""
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = re.sub(r"\s+", " ", text)
    return text


# Mots-clés de détection de colonne. Chaque entrée : (positifs, exclusions).
# La première colonne (de gauche à droite) dont l'en-tête normalisé contient
# au moins un mot-clé positif et aucune exclusion est retenue.
_NUMERO_KEYWORDS = (
    ["PHONE NUMBER", "NUMERO", "NUMÉRO", "SERVICE ID", "MSISDN", "NUMBER", "TEL"],
    ["STAFF", "MATRICULE"],  # exclusions : ne pas confondre avec un identifiant RH
)
_MONTANT_KEYWORDS = (
    ["AMOUNT", "MONTANT", "CONSOMMATION", "BILLED WITH TAX", "TTC"],
    ["PREVIOUS", "PRECEDENT", "PRÉCÉDENT"],
)
_STAFF_ID_KEYWORDS = (
    ["STAFF ID", "MATRICULE", "EMPLOYEE ID", "ID EMPLOYE", "ID COLLABORATEUR"],
    [],
)
_DATE_ACTIVATION_KEYWORDS = (
    ["DATE D'ACTIVATION", "DATE ACTIVATION", "ACTIVATION DATE"],
    [],
)
_MOIS_FACTURE_KEYWORDS = (
    ["MONTH BILLED", "MOIS FACTURE", "BILLING MONTH"],
    ["PREVIOUS", "PRECEDENT", "PRÉCÉDENT"],
)
_COMMENT_KEYWORDS = (
    ["COMMENT", "COMMENTAIRE", "REMARQUE"],
    [],
)


def _detect_column(headers: list[str], keywords: tuple[list[str], list[str]]) -> Optional[int]:
    positives, exclusions = keywords
    for idx, header in enumerate(headers):
        if any(exc in header for exc in exclusions):
            continue
        if any(pos in header for pos in positives):
            return idx
    return None


@dataclass
class ColonnesDetectees:
    numero: Optional[int]
    montant: Optional[int]
    staff_id: Optional[int]
    date_activation: Optional[int]
    mois_facture: Optional[int]
    commentaire: Optional[int]
    headers_bruts: list[str]


class FichierOperateurInvalide(Exception):
    """Levée quand un fichier opérateur ne peut pas être exploité (colonnes
    essentielles introuvables, feuille vide, etc.). Message pensé pour être
    affiché tel quel à un utilisateur non technique."""


def _choisir_feuille(wb: openpyxl.Workbook, operateur_hint: str) -> str:
    """Choisit la feuille à lire dans un classeur pouvant contenir plusieurs
    onglets (ex: 'ORANGE JUNE', 'MTN JUNE' dans un même fichier historique)."""
    noms = wb.sheetnames
    if len(noms) == 1:
        return noms[0]

    hint = normalize_header(operateur_hint)
    correspondances = [n for n in noms if hint in normalize_header(n)]
    if len(correspondances) == 1:
        return correspondances[0]
    if len(correspondances) > 1:
        # Plusieurs feuilles matchent (ex: fichier multi-mois) -> on prend
        # celle avec le plus de lignes de données, généralement la plus récente/complète.
        return max(correspondances, key=lambda n: wb[n].max_row)

    # Aucune correspondance par nom : on prend la feuille active, ou à
    # défaut celle avec le plus de lignes.
    if wb.active is not None and wb.active.max_row > 1:
        return wb.active.title
    return max(noms, key=lambda n: wb[n].max_row)


def detecter_colonnes(headers_bruts: list) -> ColonnesDetectees:
    headers = [normalize_header(h) for h in headers_bruts]
    return ColonnesDetectees(
        numero=_detect_column(headers, _NUMERO_KEYWORDS),
        montant=_detect_column(headers, _MONTANT_KEYWORDS),
        staff_id=_detect_column(headers, _STAFF_ID_KEYWORDS),
        date_activation=_detect_column(headers, _DATE_ACTIVATION_KEYWORDS),
        mois_facture=_detect_column(headers, _MOIS_FACTURE_KEYWORDS),
        commentaire=_detect_column(headers, _COMMENT_KEYWORDS),
        headers_bruts=[str(h) if h is not None else "" for h in headers_bruts],
    )


def lire_fichier_operateur(chemin: str, operateur_hint: str) -> pd.DataFrame:
    """
    Lit un fichier Excel opérateur (Orange ou MTN) et retourne un DataFrame
    à colonnes normalisées, quel que soit l'ordre/libellé des colonnes
    d'origine :

        numero_brut, montant, staff_id_brut, commentaire_brut

    `operateur_hint` sert uniquement à choisir la bonne feuille si le
    classeur en contient plusieurs (ex: "ORANGE", "MTN").

    Lève FichierOperateurInvalide avec un message clair si le fichier ne
    peut pas être exploité (colonnes numéro/montant introuvables).
    """
    try:
        wb = openpyxl.load_workbook(chemin, data_only=True, read_only=True)
    except Exception as exc:  # fichier corrompu, mauvais format, verrouillé...
        raise FichierOperateurInvalide(
            f"Impossible d'ouvrir le fichier '{chemin}' : {exc}. "
            f"Vérifiez qu'il s'agit bien d'un fichier Excel (.xlsx) non corrompu et non protégé par mot de passe."
        ) from exc

    feuille = _choisir_feuille(wb, operateur_hint)
    ws = wb[feuille]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_bruts = list(next(rows_iter))
    except StopIteration as exc:
        raise FichierOperateurInvalide(
            f"La feuille '{feuille}' du fichier '{chemin}' est vide (aucun en-tête trouvé)."
        ) from exc

    colonnes = detecter_colonnes(headers_bruts)

    if colonnes.numero is None:
        raise FichierOperateurInvalide(
            f"Impossible de trouver la colonne du numéro de téléphone dans la feuille "
            f"'{feuille}' du fichier '{chemin}'.\nColonnes trouvées : {colonnes.headers_bruts}\n"
            f"Attendu : un en-tête contenant par exemple 'NUMBER', 'NUMERO', 'SERVICE ID' ou 'PHONE'."
        )
    if colonnes.montant is None:
        raise FichierOperateurInvalide(
            f"Impossible de trouver la colonne du montant de consommation dans la feuille "
            f"'{feuille}' du fichier '{chemin}'.\nColonnes trouvées : {colonnes.headers_bruts}\n"
            f"Attendu : un en-tête contenant par exemple 'AMOUNT', 'MONTANT' ou 'TTC'."
        )

    records = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue  # ligne totalement vide -> ignorée silencieusement (pas une donnée)
        numero = row[colonnes.numero] if colonnes.numero < len(row) else None
        montant = row[colonnes.montant] if colonnes.montant < len(row) else None
        staff_id = (row[colonnes.staff_id] if colonnes.staff_id is not None
                    and colonnes.staff_id < len(row) else None)
        commentaire = (row[colonnes.commentaire] if colonnes.commentaire is not None
                       and colonnes.commentaire < len(row) else None)

        # Une ligne sans numéro n'est pas exploitable pour le rapprochement ;
        # on la conserve quand même (statut NUMERO_INVALIDE géré en aval) sauf
        # si elle est totalement vide par ailleurs (déjà filtré ci-dessus).
        records.append({
            "numero_brut": numero,
            "montant_brut": montant,
            "staff_id_brut": staff_id,
            "commentaire_brut": commentaire,
        })

    df = pd.DataFrame.from_records(
        records, columns=["numero_brut", "montant_brut", "staff_id_brut", "commentaire_brut"]
    )
    df.attrs["feuille_source"] = feuille
    df.attrs["fichier_source"] = chemin
    df.attrs["headers_bruts"] = colonnes.headers_bruts
    return df
