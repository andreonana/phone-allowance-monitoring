"""
Persistance de l'historique mensuel des consommations : `data/history.xlsx`.

Deux feuilles :
    IMPORTS        : un enregistrement par (mois, opérateur) importé
    CONSOMMATIONS  : le détail ligne à ligne (une ligne par numéro/mois/opérateur)

Le responsable ne fournit JAMAIS le fichier du mois précédent : cette
persistance s'en charge. Le mois "précédent" utilisé pour les comparaisons
est le dernier mois présent en base strictement avant le mois importé — pas
forcément M-1 calendaire, pour rester robuste si un mois a été sauté
(cahier des charges §16).

Ré-importer un mois déjà présent (ex : correction d'une erreur) REMPLACE
les données existantes pour ce couple (mois, opérateur) plutôt que de les
dupliquer : l'opération "ANALYSER" reste idempotente.
"""

from __future__ import annotations

import os
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models import LigneConsommation, LotImport, Operateur, StatutRapprochement

_IMPORTS_HEADERS = ["IMPORT_ID", "MOIS", "OPERATEUR", "NOM_FICHIER", "DATE_IMPORT",
                     "NB_LIGNES", "NB_NON_IDENTIFIES", "NB_INVALIDES"]
_CONSO_HEADERS = ["IMPORT_ID", "MOIS", "OPERATEUR", "NUMERO_BRUT", "NUMERO_NORMALISE",
                   "MONTANT", "MATRICULE", "STATUT", "STAFF_ID_SOURCE_BRUT", "COMMENTAIRE"]


class HistoryRepository:
    def __init__(self, chemin_fichier: Path):
        self.chemin_fichier = Path(chemin_fichier)
        self.imports: list[LotImport] = []
        self._import_ids: dict[tuple[str, str], str] = {}  # (mois, operateur) -> import_id
        self.consommations: list[dict] = []  # lignes brutes (dict) pour rester léger en mémoire
        self._charger()

    # ------------------------------------------------------------------ #
    def _charger(self) -> None:
        if not self.chemin_fichier.exists():
            return

        wb = openpyxl.load_workbook(self.chemin_fichier, data_only=True)

        if "IMPORTS" in wb.sheetnames:
            ws = wb["IMPORTS"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] in (None, ""):
                    continue
                import_id, mois, operateur, nom_fichier, date_import, nb_l, nb_ni, nb_inv = (list(row) + [None] * 8)[:8]
                self.imports.append(LotImport(
                    mois=str(mois), operateur=Operateur(operateur), nom_fichier=nom_fichier or "",
                    date_import=date_import if isinstance(date_import, datetime) else datetime.now(),
                    nb_lignes=int(nb_l or 0), nb_non_identifies=int(nb_ni or 0), nb_invalides=int(nb_inv or 0),
                ))
                self._import_ids[(str(mois), str(operateur))] = str(import_id)

        if "CONSOMMATIONS" in wb.sheetnames:
            ws = wb["CONSOMMATIONS"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[1] in (None, ""):  # index 1 = MOIS
                    continue
                (import_id, mois, operateur, numero_brut, numero_normalise, montant,
                 matricule, statut, staff_id_brut, commentaire) = (list(row) + [None] * 10)[:10]
                self.consommations.append({
                    "import_id": import_id, "mois": str(mois), "operateur": operateur,
                    "numero_brut": numero_brut, "numero_normalise": numero_normalise,
                    "montant": float(montant) if montant is not None else 0.0,
                    "matricule": matricule, "statut": statut,
                    "staff_id_source_brut": staff_id_brut, "commentaire": commentaire or "",
                })

    def save(self) -> None:
        self.chemin_fichier.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_i: Worksheet = wb.create_sheet("IMPORTS")
        ws_i.append(_IMPORTS_HEADERS)
        for (mois, operateur), import_id in sorted(self._import_ids.items()):
            lot = next((i for i in self.imports if i.mois == mois and i.operateur.value == operateur), None)
            if lot is None:
                continue
            ws_i.append([import_id, lot.mois, lot.operateur.value, lot.nom_fichier, lot.date_import,
                         lot.nb_lignes, lot.nb_non_identifies, lot.nb_invalides])

        ws_c: Worksheet = wb.create_sheet("CONSOMMATIONS")
        ws_c.append(_CONSO_HEADERS)
        for c in self.consommations:
            ws_c.append([c["import_id"], c["mois"], c["operateur"], c["numero_brut"],
                         c["numero_normalise"], c["montant"], c["matricule"], c["statut"],
                         c["staff_id_source_brut"], c["commentaire"]])

        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=str(self.chemin_fichier.parent))
        os.close(fd)
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, self.chemin_fichier)
        except Exception:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            raise

    # ------------------------------------------------------------------ #
    # Écriture d'un lot mensuel (remplace tout lot existant pour le même
    # couple mois/opérateur -> idempotence en cas de ré-analyse).
    # ------------------------------------------------------------------ #

    def enregistrer_lot(self, mois: str, operateur: Operateur, nom_fichier: str,
                         lignes: list[LigneConsommation]) -> LotImport:
        cle = (mois, operateur.value)
        # Purge de l'éventuel lot précédent pour ce mois/opérateur
        ancien_id = self._import_ids.get(cle)
        if ancien_id is not None:
            self.consommations = [c for c in self.consommations if c["import_id"] != ancien_id]
            self.imports = [i for i in self.imports if not (i.mois == mois and i.operateur == operateur)]

        import_id = str(uuid.uuid4())
        self._import_ids[cle] = import_id

        nb_non_id = sum(1 for l in lignes if l.statut == StatutRapprochement.NON_IDENTIFIE)
        nb_inval = sum(1 for l in lignes if l.statut == StatutRapprochement.NUMERO_INVALIDE)
        lot = LotImport(mois=mois, operateur=operateur, nom_fichier=nom_fichier, date_import=datetime.now(),
                         nb_lignes=len(lignes), nb_non_identifies=nb_non_id, nb_invalides=nb_inval)
        self.imports.append(lot)

        for l in lignes:
            self.consommations.append({
                "import_id": import_id, "mois": l.mois, "operateur": l.operateur.value,
                "numero_brut": l.numero_brut, "numero_normalise": l.numero_normalise,
                "montant": l.montant, "matricule": l.matricule, "statut": l.statut.value,
                "staff_id_source_brut": l.staff_id_source_brut, "commentaire": l.commentaire,
            })
        return lot

    # ------------------------------------------------------------------ #
    # Lecture / requêtes
    # ------------------------------------------------------------------ #

    def mois_disponibles(self) -> list[str]:
        return sorted({c["mois"] for c in self.consommations})

    def dernier_mois(self) -> Optional[str]:
        mois = self.mois_disponibles()
        return mois[-1] if mois else None

    def mois_precedent(self, mois: str) -> Optional[str]:
        """Dernier mois présent en base strictement avant `mois` (pas
        forcément M-1 calendaire)."""
        anterieurs = [m for m in self.mois_disponibles() if m < mois]
        return max(anterieurs) if anterieurs else None

    def lignes_du_mois(self, mois: str, operateur: Optional[Operateur] = None) -> list[dict]:
        return [c for c in self.consommations
                if c["mois"] == mois and (operateur is None or c["operateur"] == operateur.value)]

    def totaux_par_matricule(self, mois: str) -> dict[str, dict]:
        """
        Retourne {matricule: {"ORANGE": montant, "MTN": montant, "total": montant,
                               "numero_orange": str, "numero_mtn": str}}
        pour un mois donné. Les numéros NON_IDENTIFIE (matricule=None) sont exclus
        de cet agrégat (traités séparément par le moteur d'alertes).
        """
        resultat: dict[str, dict] = {}
        for c in self.lignes_du_mois(mois):
            if not c["matricule"] or c["statut"] != StatutRapprochement.OK.value:
                continue
            entry = resultat.setdefault(c["matricule"], {
                "ORANGE": 0.0, "MTN": 0.0, "total": 0.0, "numero_orange": "", "numero_mtn": "",
            })
            entry[c["operateur"]] += c["montant"]
            entry["total"] += c["montant"]
            if c["operateur"] == Operateur.ORANGE.value:
                entry["numero_orange"] = c["numero_normalise"] or entry["numero_orange"]
            else:
                entry["numero_mtn"] = c["numero_normalise"] or entry["numero_mtn"]
        return resultat

    def totaux_globaux_par_mois(self) -> dict[str, dict]:
        """
        Retourne {mois: {"ORANGE": montant, "MTN": montant, "total": montant}}
        agrégé sur TOUTES les lignes du mois (y compris non identifiées et
        invalides, qui restent des montants réellement facturés) — sert à
        tracer l'évolution globale de la consommation dans le temps
        (tableau de bord), indépendamment du rattachement à un matricule.
        """
        resultat: dict[str, dict] = {}
        for c in self.consommations:
            entry = resultat.setdefault(c["mois"], {"ORANGE": 0.0, "MTN": 0.0, "total": 0.0})
            entry[c["operateur"]] += c["montant"]
            entry["total"] += c["montant"]
        return resultat

    def numeros_du_mois(self, mois: str, operateur: Operateur) -> set[str]:
        return {c["numero_normalise"] for c in self.lignes_du_mois(mois, operateur)
                if c["numero_normalise"]}
