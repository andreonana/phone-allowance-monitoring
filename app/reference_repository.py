"""
Persistance de la base de référence interne : NUMÉRO <-> MATRICULE <-> COLLABORATEUR.

Aucune base de données : le classeur `data/reference_base.xlsx` est LA
source de vérité, avec deux feuilles :

    COLLABORATEURS : MATRICULE, NOM, DIRECTION, FONCTION, TYPE, ACTIF, ...
    NUMEROS        : NUMERO_NORMALISE, OPERATEUR, MATRICULE, ACTIF, ...

Cette base est créée automatiquement au premier lancement (vide), puis :
    - alimentée une première fois via `bootstrap_from_operator_file()` à
      partir d'un fichier opérateur qui contient déjà une colonne
      STAFF ID / MATRICULE (ex : le classeur DRAFT fourni) ;
    - complétée/corrigée ensuite manuellement dans l'application (ajout,
      modification, suppression de collaborateurs et de numéros) ;
    - PLUS JAMAIS réimportée automatiquement chaque mois (cahier des
      charges §4) : les fichiers Orange/MTN mensuels "normaux" ne
      contiennent pas de matricule et n'ont pas besoin d'en contenir.

Écriture atomique : on écrit dans un fichier temporaire puis on remplace
l'ancien fichier, pour ne jamais laisser un classeur à moitié écrit en cas
de coupure/erreur pendant la sauvegarde.
"""

from __future__ import annotations

import os
import re
import tempfile
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Collaborateur, NumeroReference, Operateur, TypeCollaborateur
from app.normalization import normalize_phone_number

_COLLAB_HEADERS = ["MATRICULE", "NOM", "DIRECTION", "FONCTION", "TYPE", "ACTIF",
                    "DATE_CREATION", "DATE_MODIFICATION"]
_NUMERO_HEADERS = ["NUMERO_NORMALISE", "OPERATEUR", "MATRICULE", "ACTIF", "DATE_CREATION"]

# Un matricule "standard" ressemble à CM0814, OS0244, IH0020, A28234, BF100548...
# (1 à 4 lettres suivies de chiffres). Tout le reste (nom de service, phrase
# libre) est considéré comme un compte de type POOL par défaut.
_CODE_MATRICULE_PATTERN = re.compile(r"^[A-Z]{1,4}\d{3,7}$")


def _guess_type(matricule: str) -> TypeCollaborateur:
    return (TypeCollaborateur.INDIVIDUEL
            if _CODE_MATRICULE_PATTERN.match(matricule.strip().upper())
            else TypeCollaborateur.POOL)


class ReferenceRepository:
    """Charge/persiste la base de référence en mémoire, avec repli sûr en
    cas de fichier absent ou vide."""

    def __init__(self, chemin_fichier: Path):
        self.chemin_fichier = Path(chemin_fichier)
        self.collaborateurs: dict[str, Collaborateur] = {}
        self.numeros: dict[str, NumeroReference] = {}  # clé = numero_normalise
        self._charger()

    # ------------------------------------------------------------------ #
    # Chargement / sauvegarde
    # ------------------------------------------------------------------ #

    def _charger(self) -> None:
        if not self.chemin_fichier.exists():
            return  # base vide, sera créée à la première sauvegarde

        wb = openpyxl.load_workbook(self.chemin_fichier, data_only=True)

        if "COLLABORATEURS" in wb.sheetnames:
            ws = wb["COLLABORATEURS"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] in (None, ""):
                    continue
                matricule, nom, direction, fonction, type_, actif, dc, dm = (list(row) + [None] * 8)[:8]
                self.collaborateurs[str(matricule).strip()] = Collaborateur(
                    matricule=str(matricule).strip(),
                    nom=nom or "",
                    direction=direction or "",
                    fonction=fonction or "",
                    type=TypeCollaborateur(type_) if type_ in TypeCollaborateur._value2member_map_ else TypeCollaborateur.INDIVIDUEL,
                    actif=bool(actif) if actif is not None else True,
                    date_creation=dc if isinstance(dc, datetime) else datetime.now(),
                    date_modification=dm if isinstance(dm, datetime) else datetime.now(),
                )

        if "NUMEROS" in wb.sheetnames:
            ws = wb["NUMEROS"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] in (None, ""):
                    continue
                numero, operateur, matricule, actif, dc = (list(row) + [None] * 5)[:5]
                numero = str(numero).strip()
                self.numeros[numero] = NumeroReference(
                    numero_normalise=numero,
                    operateur=Operateur(operateur) if operateur in Operateur._value2member_map_ else Operateur.ORANGE,
                    matricule=str(matricule).strip() if matricule else "",
                    actif=bool(actif) if actif is not None else True,
                    date_creation=dc if isinstance(dc, datetime) else datetime.now(),
                )

    def save(self) -> None:
        """Écriture atomique du classeur de référence."""
        self.chemin_fichier.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_c: Worksheet = wb.create_sheet("COLLABORATEURS")
        ws_c.append(_COLLAB_HEADERS)
        for c in sorted(self.collaborateurs.values(), key=lambda c: c.matricule):
            ws_c.append([c.matricule, c.nom, c.direction, c.fonction, c.type.value,
                         c.actif, c.date_creation, c.date_modification])

        ws_n: Worksheet = wb.create_sheet("NUMEROS")
        ws_n.append(_NUMERO_HEADERS)
        for n in sorted(self.numeros.values(), key=lambda n: n.numero_normalise):
            ws_n.append([n.numero_normalise, n.operateur.value, n.matricule, n.actif, n.date_creation])

        for ws in (ws_c, ws_n):
            for col_cells in ws.columns:
                width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 12), 40)

        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=str(self.chemin_fichier.parent))
        os.close(fd)
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, self.chemin_fichier)  # remplacement atomique
        except Exception:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            raise

    # ------------------------------------------------------------------ #
    # Lecture
    # ------------------------------------------------------------------ #

    def get_matricule_pour_numero(self, numero_normalise: str) -> Optional[str]:
        ref = self.numeros.get(numero_normalise)
        if ref is None or not ref.actif:
            return None
        return ref.matricule

    def get_collaborateur(self, matricule: str) -> Optional[Collaborateur]:
        return self.collaborateurs.get(matricule)

    def numeros_actifs_du_matricule(self, matricule: str, operateur: Optional[Operateur] = None) -> list[NumeroReference]:
        return [n for n in self.numeros.values()
                if n.matricule == matricule and n.actif and (operateur is None or n.operateur == operateur)]

    # ------------------------------------------------------------------ #
    # Écriture / édition (ajout, modification, suppression)
    # ------------------------------------------------------------------ #

    def upsert_collaborateur(self, matricule: str, nom: str = "", direction: str = "",
                              fonction: str = "", type: Optional[TypeCollaborateur] = None) -> Collaborateur:
        matricule = matricule.strip()
        existant = self.collaborateurs.get(matricule)
        collab = Collaborateur(
            matricule=matricule,
            nom=nom or (existant.nom if existant else ""),
            direction=direction or (existant.direction if existant else ""),
            fonction=fonction or (existant.fonction if existant else ""),
            type=type or (existant.type if existant else _guess_type(matricule)),
            actif=existant.actif if existant else True,
            date_creation=existant.date_creation if existant else datetime.now(),
            date_modification=datetime.now(),
        )
        self.collaborateurs[matricule] = collab
        return collab

    def supprimer_collaborateur(self, matricule: str, *, supprimer_numeros: bool = True) -> None:
        self.collaborateurs.pop(matricule, None)
        if supprimer_numeros:
            for num in [n for n, ref in self.numeros.items() if ref.matricule == matricule]:
                del self.numeros[num]

    def upsert_numero(self, numero_brut, operateur: Operateur, matricule: str) -> Optional[NumeroReference]:
        """Ajoute/actualise un numéro dans la base après normalisation.
        Retourne None (et n'écrit rien) si le numéro brut est invalide."""
        resultat = normalize_phone_number(numero_brut)
        if not resultat.is_valid:
            return None
        ref = NumeroReference(
            numero_normalise=resultat.normalized,
            operateur=operateur,
            matricule=matricule.strip(),
            actif=True,
            date_creation=self.numeros.get(resultat.normalized).date_creation
            if resultat.normalized in self.numeros else datetime.now(),
        )
        self.numeros[resultat.normalized] = ref
        return ref

    def supprimer_numero(self, numero_normalise: str) -> None:
        self.numeros.pop(numero_normalise, None)

    # ------------------------------------------------------------------ #
    # Bootstrap depuis un fichier opérateur contenant déjà un matricule
    # ------------------------------------------------------------------ #

    def bootstrap_from_operator_file(self, chemin: str, operateur: Operateur) -> dict:
        """
        Alimente la base de référence à partir d'un fichier opérateur qui
        contient une colonne matricule/STAFF ID déjà renseignée (ex : le
        classeur DRAFT fourni par le client, résultat du rapprochement
        manuel actuel). Usage ponctuel (création initiale / mise à jour de
        la base), PAS le flux mensuel normal.

        Règles de nettoyage appliquées (cf. analyse des fichiers réels) :
            - espaces de tête/fin supprimés du matricule ;
            - valeurs '#N/A' ou vides -> numéro ignoré (reste NON IDENTIFIE,
              à corriger manuellement, jamais inventé automatiquement) ;
            - matricule au format standard (ex: CM0814) -> collaborateur
              INDIVIDUEL ; tout autre libellé (nom de service, phrase) ->
              compte POOL, conformément au choix retenu pour les lignes
              partagées (agences, projets...).

        Retourne un résumé {nb_numeros_ajoutes, nb_ignores_non_identifie,
        nb_collaborateurs_crees}.
        """
        from app.excel_io import lire_fichier_operateur  # import local : évite un cycle

        df = lire_fichier_operateur(chemin, operateur.value)
        nb_ajoutes = 0
        nb_ignores = 0
        matricules_crees: set[str] = set()

        for _, row in df.iterrows():
            staff_id_brut = row["staff_id_brut"]
            if staff_id_brut is None:
                nb_ignores += 1
                continue
            matricule = str(staff_id_brut).strip()
            if not matricule or matricule.upper() in ("#N/A", "N/A", "NA", "NONE"):
                nb_ignores += 1
                continue

            if matricule not in self.collaborateurs:
                self.upsert_collaborateur(matricule)
                matricules_crees.add(matricule)

            ref = self.upsert_numero(row["numero_brut"], operateur, matricule)
            if ref is not None:
                nb_ajoutes += 1
            else:
                nb_ignores += 1  # numéro brut invalide, non intégré à la base

        return {
            "nb_numeros_ajoutes": nb_ajoutes,
            "nb_ignores": nb_ignores,
            "nb_collaborateurs_crees": len(matricules_crees),
        }
