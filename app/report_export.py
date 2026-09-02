"""
Génération du rapport Excel final multi-feuilles pour un mois donné :

    SYNTHESE  : une ligne par collaborateur (Orange + MTN consolidés,
                comparaison au mois précédent, alertes associées)
    ALERTES   : toutes les alertes du mois, triées par type
    DETAIL    : le détail brut de toutes les lignes de consommation importées
                (y compris invalides / non identifiées, jamais filtrées — §7)

Écrit dans `data/exports/` (cf. `app.paths.get_exports_dir`), un fichier par
export (nom horodaté) : un export ne modifie ni ne remplace jamais les
données persistées (`history.xlsx`, `reference_base.xlsx`).
"""

from __future__ import annotations

import os
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.history_repository import HistoryRepository
from app.models import Alerte, SyntheseCollaborateur
from app.paths import get_exports_dir

_ENTETE_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
_ENTETE_FONT = Font(color="FFFFFFFF", bold=True)

# Code couleur partagé entre le STATUT de la synthèse et le TYPE des alertes :
# fond clair + texte de la même teinte foncée, pour rester lisible tout en
# restant reconnaissable d'un coup d'œil (rouge = ça coûte plus cher / à
# surveiller, vert = ça baisse, gris = rien à signaler, bleu = nouveauté,
# ambre/orange = donnée à corriger dans la base de référence).
_COULEURS_STATUT = {
    "Hausse": ("FFFEE2E2", "FF991B1B"),
    "Baisse": ("FFDCFCE7", "FF166534"),
    "Stable": ("FFF3F4F6", "FF374151"),
    "Nouveau": ("FFDBEAFE", "FF1E40AF"),
    "Disparu": ("FFFEF3C7", "FF92400E"),
}
_COULEURS_ALERTE = {
    "FORTE_AUGMENTATION": ("FFFEE2E2", "FF991B1B"),
    "FORTE_CONSOMMATION": ("FFFFEDD5", "FF9A3412"),
    "NUMERO_NON_IDENTIFIE": ("FFFEF3C7", "FF92400E"),
    "NUMERO_INVALIDE": ("FFF3F4F6", "FF374151"),
    "NOUVEAU_NUMERO": ("FFDBEAFE", "FF1E40AF"),
    "NUMERO_DISPARU": ("FFEDE9FE", "FF5B21B6"),
    "DOUBLON": ("FFFEE2E2", "FF7F1D1D"),
}


def _colorer_cellule(ws: Worksheet, ligne: int, colonne: int, couleurs: tuple[str, str]) -> None:
    fond, texte = couleurs
    cellule = ws.cell(row=ligne, column=colonne)
    cellule.fill = PatternFill(start_color=fond, end_color=fond, fill_type="solid")
    cellule.font = Font(color=texte, bold=True)

_SYNTHESE_HEADERS = [
    "MATRICULE", "NOM", "DIRECTION", "FONCTION", "TYPE",
    "NUMERO ORANGE", "MONTANT ORANGE", "NUMERO MTN", "MONTANT MTN",
    "TOTAL MOIS ACTUEL", "TOTAL MOIS PRECEDENT", "VARIATION (FCFA)",
    "VARIATION (%)", "STATUT", "ALERTES",
]
_ALERTES_HEADERS = ["TYPE", "MATRICULE", "NOM", "NUMERO", "MESSAGE", "VALEUR", "SEUIL APPLIQUE"]
_DETAIL_HEADERS = ["MOIS", "OPERATEUR", "NUMERO_BRUT", "NUMERO_NORMALISE", "MONTANT",
                    "MATRICULE", "STATUT", "STAFF_ID_SOURCE_BRUT", "COMMENTAIRE"]


def _styler_entete(ws: Worksheet, nb_colonnes: int) -> None:
    for col in range(1, nb_colonnes + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _ENTETE_FILL
        cell.font = _ENTETE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        largeur = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(largeur + 2, 10), 50)


_COL_STATUT_SYNTHESE = _SYNTHESE_HEADERS.index("STATUT") + 1
_COL_TYPE_ALERTE = _ALERTES_HEADERS.index("TYPE") + 1


def _feuille_synthese(wb: openpyxl.Workbook, syntheses: list[SyntheseCollaborateur]) -> None:
    ws = wb.create_sheet("SYNTHESE")
    ws.append(_SYNTHESE_HEADERS)
    for ligne, s in enumerate(sorted(syntheses, key=lambda s: s.total_actuel, reverse=True), start=2):
        ws.append([
            s.matricule, s.nom, s.direction, s.fonction, s.type.value,
            s.numero_orange, s.montant_orange, s.numero_mtn, s.montant_mtn,
            s.total_actuel, s.total_precedent, s.variation_montant, s.variation_pct,
            s.statut, "; ".join(a.type.value for a in s.alertes),
        ])
        couleurs = _COULEURS_STATUT.get(s.statut)
        if couleurs:
            _colorer_cellule(ws, ligne, _COL_STATUT_SYNTHESE, couleurs)
    _styler_entete(ws, len(_SYNTHESE_HEADERS))
    _autosize(ws)


def _feuille_alertes(wb: openpyxl.Workbook, alertes: list[Alerte]) -> None:
    ws = wb.create_sheet("ALERTES")
    ws.append(_ALERTES_HEADERS)
    for ligne, a in enumerate(sorted(alertes, key=lambda a: a.type.value), start=2):
        ws.append([a.type.value, a.matricule or "", a.nom, a.numero, a.message,
                   a.valeur, a.seuil_applique])
        couleurs = _COULEURS_ALERTE.get(a.type.value)
        if couleurs:
            _colorer_cellule(ws, ligne, _COL_TYPE_ALERTE, couleurs)
    _styler_entete(ws, len(_ALERTES_HEADERS))
    _autosize(ws)


def _feuille_detail(wb: openpyxl.Workbook, mois: str, history: HistoryRepository) -> None:
    ws = wb.create_sheet("DETAIL")
    ws.append(_DETAIL_HEADERS)
    for c in history.lignes_du_mois(mois):
        ws.append([c["mois"], c["operateur"], c["numero_brut"], c["numero_normalise"],
                   c["montant"], c["matricule"] or "", c["statut"],
                   c["staff_id_source_brut"], c["commentaire"]])
    _styler_entete(ws, len(_DETAIL_HEADERS))
    _autosize(ws)


def exporter_rapport(mois: str, syntheses: list[SyntheseCollaborateur], alertes: list[Alerte],
                      history: HistoryRepository, *, dossier: Path | None = None) -> Path:
    """
    Génère le classeur de rapport du mois et retourne le chemin du fichier
    créé. `dossier` est surtout utile pour les tests (répertoire temporaire) ;
    en usage normal, laisser `app.paths.get_exports_dir()` par défaut.
    """
    dossier = Path(dossier) if dossier is not None else get_exports_dir()
    dossier.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _feuille_synthese(wb, syntheses)
    _feuille_alertes(wb, alertes)
    _feuille_detail(wb, mois, history)

    horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
    chemin_final = dossier / f"rapport_{mois}_{horodatage}.xlsx"

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=str(dossier))
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, chemin_final)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise

    return chemin_final
