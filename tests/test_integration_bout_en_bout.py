"""Test d'intégration bout-en-bout : fichiers Excel opérateurs -> lecture ->
rapprochement -> historique -> consolidation -> alertes -> export du rapport.
Reproduit le flux réellement exécuté par l'IHM (`ui.app_window`)."""

import openpyxl

from app.alerts_engine import generer_alertes
from app.config import ConfigRepository
from app.consolidation import consolider
from app.excel_io import lire_fichier_operateur
from app.history_repository import HistoryRepository
from app.models import Operateur
from app.reconciliation_engine import rapprocher_fichier
from app.reference_repository import ReferenceRepository
from app.report_export import exporter_rapport


def _creer_fichier_orange(chemin):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ORANGE JUNE"
    ws.append(["Service ID", "Amount Billed with tax"])
    ws.append(["699 123 456", "12 000"])   # CM0814, connu
    ws.append(["+237677000000", 5000])      # inconnu de la base -> NON_IDENTIFIE
    ws.append(["notanumber", 1000])         # invalide -> NUMERO_INVALIDE
    wb.save(chemin)


def _creer_fichier_mtn(chemin):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MTN JUNE"
    ws.append(["Number", "Montant"])
    ws.append(["670987654", 8000])  # CM0814 aussi côté MTN
    wb.save(chemin)


def test_flux_complet(tmp_path):
    fichier_orange = tmp_path / "orange.xlsx"
    fichier_mtn = tmp_path / "mtn.xlsx"
    _creer_fichier_orange(fichier_orange)
    _creer_fichier_mtn(fichier_mtn)

    reference = ReferenceRepository(tmp_path / "reference_base.xlsx")
    reference.upsert_collaborateur("CM0814", nom="Jean Dupont", direction="IT")
    reference.upsert_numero("699123456", Operateur.ORANGE, "CM0814")
    reference.upsert_numero("670987654", Operateur.MTN, "CM0814")
    reference.save()

    history = HistoryRepository(tmp_path / "history.xlsx")
    config = ConfigRepository(tmp_path / "config.xlsx")
    config.set("SEUIL_FORTE_CONSOMMATION_FCFA", 15000.0)

    for operateur, chemin in ((Operateur.ORANGE, fichier_orange), (Operateur.MTN, fichier_mtn)):
        df = lire_fichier_operateur(str(chemin), operateur.value)
        lignes = rapprocher_fichier(df, "2026-06", operateur, reference)
        history.enregistrer_lot("2026-06", operateur, chemin.name, lignes)
    history.save()

    # Rapprochement : CM0814 rattaché des deux côtés, un numéro non identifié, un invalide.
    lignes = history.lignes_du_mois("2026-06")
    assert len(lignes) == 4  # 3 Orange + 1 MTN, aucune ligne perdue (§7)

    syntheses = consolider("2026-06", reference, history)
    assert len(syntheses) == 1
    s = syntheses[0]
    assert s.matricule == "CM0814"
    assert s.montant_orange == 12000
    assert s.montant_mtn == 8000
    assert s.total_actuel == 20000
    assert s.statut == "Nouveau"  # pas de mois précédent en base

    alertes = generer_alertes("2026-06", syntheses, history, reference, config)
    types = {a.type.value for a in alertes}
    assert "NUMERO_NON_IDENTIFIE" in types
    assert "NUMERO_INVALIDE" in types
    assert "FORTE_CONSOMMATION" in types  # 20000 >= seuil 15000

    chemin_rapport = exporter_rapport("2026-06", syntheses, alertes, history, dossier=tmp_path / "exports")
    assert chemin_rapport.exists()

    wb = openpyxl.load_workbook(chemin_rapport)
    ligne_synthese = list(wb["SYNTHESE"].iter_rows(min_row=2, values_only=True))[0]
    assert ligne_synthese[0] == "CM0814"
    assert ligne_synthese[9] == 20000  # TOTAL MOIS ACTUEL
