import openpyxl

from app.alerts_engine import generer_alertes
from app.config import ConfigRepository
from app.consolidation import consolider
from app.history_repository import HistoryRepository
from app.models import LigneConsommation, Operateur, StatutRapprochement
from app.reference_repository import ReferenceRepository
from app.report_export import exporter_rapport


def test_export_genere_les_trois_feuilles(tmp_path):
    reference = ReferenceRepository(tmp_path / "reference_base.xlsx")
    reference.upsert_collaborateur("CM0814", nom="Jean Dupont")
    history = HistoryRepository(tmp_path / "history.xlsx")
    history.enregistrer_lot("2026-08", Operateur.ORANGE, "aout.xlsx", [
        LigneConsommation(mois="2026-08", operateur=Operateur.ORANGE, numero_brut="699123456",
                           numero_normalise="699123456", montant=150000, matricule="CM0814",
                           statut=StatutRapprochement.OK),
    ])
    config = ConfigRepository(tmp_path / "config.xlsx")

    syntheses = consolider("2026-08", reference, history)
    alertes = generer_alertes("2026-08", syntheses, history, reference, config)

    chemin = exporter_rapport("2026-08", syntheses, alertes, history, dossier=tmp_path / "exports")

    assert chemin.exists()
    wb = openpyxl.load_workbook(chemin)
    assert set(wb.sheetnames) == {"SYNTHESE", "ALERTES", "DETAIL"}

    ws = wb["SYNTHESE"]
    ligne = list(ws.iter_rows(min_row=2, values_only=True))[0]
    assert ligne[0] == "CM0814"
    assert ligne[1] == "Jean Dupont"

    assert wb["ALERTES"].max_row >= 2  # au moins l'alerte forte consommation
